#!/usr/bin/env python3
"""
ICS password crawler + Teams self-chat sender for Windows.

Run this file from C:\\Python_scripts\\ICS with the local .venv Python:

    .venv\\Scripts\\python.exe ICS_send_PW_v7.py --sso-id yhchoi20 --teams-user yhchoi20@lginnotek.com

Credential lookup order:

1. ICS_SSO_ID + ICS_SSO_PW environment variables
2. keyring service ICS_SSO, user from --sso-id or ICS_SSO_ID
3. BMS_budget_ini.xlsx next to this script

This does not use Microsoft Graph or Teams Workflows. It opens a Microsoft Teams
chat deep link, pastes the crawled password into the compose box, then sends a
Teams send key sequence while the compose box still has focus.
"""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
import time
import urllib.parse
from pathlib import Path


PORTAL_URL = "http://gportal.lginnotek.com/portal/main/portalMain.do#"
ICS_SSO_URL = "https://ics.lginnotek.com:48011/sso/checkIcsLogin.jsp"
ICS_COOKIE_NAME = "icsPwd"

DEFAULT_CONFIG_EXCEL = "BMS_budget_ini.xlsx"
DEFAULT_KEYRING_SERVICE = "ICS_SSO"
DEFAULT_MAX_URL_CHARS = 12_000
DEFAULT_SEND_DELAY = 15.0
DEFAULT_SEND_KEY = "auto"
DEFAULT_SEND_METHOD = "keys"
DEFAULT_TEAMS_READY_TIMEOUT = 30.0
DEFAULT_TEAMS_INPUT_METHOD = "clipboard"
DEFAULT_COOKIE_TIMEOUT = 30.0
DEFAULT_LOGIN_WAIT = 2.0
DEFAULT_ICS_RETRIES = 4
DEFAULT_ICS_RETRY_DELAY = 4.0


def read_sso_credentials(
    *,
    config_excel: str,
    sso_id: str | None,
    keyring_service: str,
) -> tuple[str, str]:
    env_id = sso_id or os.environ.get("ICS_SSO_ID")
    env_pw = os.environ.get("ICS_SSO_PW")
    if env_id and env_pw:
        return env_id, env_pw

    if env_id:
        keyring_pw = read_sso_password_from_keyring(keyring_service, env_id)
        if keyring_pw:
            return env_id, keyring_pw

    excel_path = Path(config_excel)
    if not excel_path.exists():
        raise FileNotFoundError(
            "SSO credentials were not found. Set ICS_SSO_ID and ICS_SSO_PW, "
            f"store the password in keyring service '{keyring_service}' for the SSO ID, "
            f"or place {config_excel} next to this script."
        )

    try:
        return read_sso_credentials_with_openpyxl(excel_path)
    except ImportError:
        return read_sso_credentials_with_pandas(excel_path)


def read_sso_password_from_keyring(service: str, user: str) -> str | None:
    try:
        import keyring
    except ImportError:
        return None

    password = keyring.get_password(service, user)
    if password:
        return password.strip()
    return None


def read_sso_credentials_with_openpyxl(excel_path: Path) -> tuple[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is not installed") from exc

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    content_col = None
    start_row = 0
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            if str(value).strip() == "내용":
                content_col = col_index
                start_row = row_index + 1
                break
        if content_col is not None:
            break

    if content_col is None:
        content_col = 1

    values: dict[str, str] = {}
    for row in rows[start_row:]:
        if not row:
            continue
        key = str(row[0]).strip() if row[0] is not None else ""
        if key in {"SSO ID", "SSO PW"} and len(row) > content_col and row[content_col] is not None:
            values[key] = str(row[content_col])

    if "SSO ID" not in values or "SSO PW" not in values:
        raise ValueError(f"Could not find 'SSO ID' and 'SSO PW' rows in {excel_path}.")

    return values["SSO ID"], values["SSO PW"]


def read_sso_credentials_with_pandas(excel_path: Path) -> tuple[str, str]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError(
            "Neither openpyxl nor pandas is installed. Set ICS_SSO_ID and ICS_SSO_PW instead."
        ) from exc

    df = pd.read_excel(excel_path, index_col=0)
    return str(df.loc["SSO ID", "내용"]), str(df.loc["SSO PW", "내용"])


def fast_wait_for_cookie(driver, name: str, timeout: float) -> dict:
    end_time = time.time() + timeout
    while time.time() < end_time:
        cookie = driver.get_cookie(name)
        if cookie:
            return cookie
        time.sleep(0.1)
    raise TimeoutError(f"Cookie '{name}' not found within {timeout} seconds.")


def wait_for_document_ready(driver, timeout: float) -> None:
    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            ready_state = driver.execute_script("return document.readyState")
            if ready_state in {"interactive", "complete"}:
                return
        except Exception:
            pass
        time.sleep(0.2)


def set_input_value(driver, element, value: str) -> None:
    driver.execute_script(
        """
        const element = arguments[0];
        const value = arguments[1];
        element.focus();
        element.value = value;
        element.dispatchEvent(new Event('input', { bubbles: true }));
        element.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        element,
        value,
    )


def save_debug_snapshot(driver, debug_dir: str | None, label: str) -> None:
    if not debug_dir:
        return

    debug_path = Path(debug_dir)
    debug_path.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in label)
    meta_file = debug_path / f"{stamp}_{safe_label}.txt"
    screenshot_file = debug_path / f"{stamp}_{safe_label}.png"

    try:
        current_url = driver.current_url
    except Exception as exc:
        current_url = f"<unavailable: {exc}>"

    try:
        title = driver.title
    except Exception as exc:
        title = f"<unavailable: {exc}>"

    meta_file.write_text(
        f"url={current_url}\ntitle={title}\n",
        encoding="utf-8",
    )

    try:
        driver.save_screenshot(str(screenshot_file))
    except Exception:
        pass


def make_browser_options(browser: str, headless: bool, detach: bool):
    if browser == "edge":
        from selenium.webdriver.edge.options import Options
    elif browser == "chrome":
        from selenium.webdriver.chrome.options import Options
    else:
        raise ValueError(f"Unsupported browser: {browser}")

    options = Options()

    try:
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
    except Exception:
        pass

    options.add_argument("--disable-features=LocalNetworkAccessChecks")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1280,900")

    if headless:
        options.add_argument("--headless=new")

    if detach:
        try:
            options.add_experimental_option("detach", True)
        except Exception:
            pass

    binary_env = os.environ.get("ICS_BROWSER_BINARY")
    if binary_env:
        options.binary_location = binary_env

    return options


def make_webdriver(browser: str, headless: bool, detach: bool):
    try:
        from selenium import webdriver
    except ImportError as exc:
        raise ImportError("selenium is required in .venv: .venv\\Scripts\\python.exe -m pip install selenium") from exc

    options = make_browser_options(browser, headless=headless, detach=detach)

    if browser == "edge":
        return webdriver.Edge(options=options)
    if browser == "chrome":
        return webdriver.Chrome(options=options)

    raise ValueError(f"Unsupported browser: {browser}")


def choose_browser(preferred: str) -> str:
    if preferred in {"chrome", "edge"}:
        return preferred

    if shutil.which("msedge") or shutil.which("microsoft-edge"):
        return "edge"
    if shutil.which("chrome") or shutil.which("google-chrome"):
        return "chrome"

    return "edge"


def crawl_ics_password(
    *,
    sso_id: str,
    sso_pw: str,
    browser: str,
    headless: bool,
    keep_browser_open: bool,
    cookie_timeout: float,
    login_wait: float,
    ics_retries: int,
    ics_retry_delay: float,
    debug_dir: str | None,
) -> str:
    from selenium.webdriver.common.by import By

    driver = make_webdriver(browser, headless=headless, detach=keep_browser_open)

    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {
                "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                })
                """
            },
        )

        driver.get(PORTAL_URL)
        driver.implicitly_wait(3)
        wait_for_document_ready(driver, timeout=10)

        input_id = driver.find_element(By.CSS_SELECTOR, 'input[name="userid"]')
        input_id.click()
        set_input_value(driver, input_id, sso_id)

        input_pw = driver.find_element(By.CSS_SELECTOR, "input#password")
        input_pw.click()
        set_input_value(driver, input_pw, sso_pw)

        login_btn = driver.find_element(By.CSS_SELECTOR, "input[type=image]")
        login_btn.click()
        time.sleep(login_wait)

        last_error: Exception | None = None
        for attempt in range(1, ics_retries + 1):
            try:
                driver.get(ICS_SSO_URL)
                wait_for_document_ready(driver, timeout=10)
                ics_cookie = fast_wait_for_cookie(driver, ICS_COOKIE_NAME, timeout=cookie_timeout)
                return str(ics_cookie["value"])
            except Exception as exc:
                last_error = exc
                print(
                    f"ICS SSO attempt {attempt}/{ics_retries} failed: {exc}",
                    file=sys.stderr,
                )
                save_debug_snapshot(driver, debug_dir, f"ics_sso_attempt_{attempt}")
                if attempt < ics_retries:
                    time.sleep(ics_retry_delay)

        assert last_error is not None
        raise last_error
    finally:
        if not keep_browser_open:
            driver.quit()


def build_teams_chat_deep_link(
    *,
    user: str,
    message: str | None = None,
    tenant_id: str | None = None,
) -> str:
    params = {"users": user}
    if message is not None:
        params["message"] = message
    if tenant_id:
        params["tenantId"] = tenant_id

    return "https://teams.microsoft.com/l/chat/0/0?" + urllib.parse.urlencode(params)


def open_url(url: str) -> None:
    startfile = getattr(os, "startfile", None)
    if startfile:
        startfile(url)
        return

    commands = [
        ["cmd.exe", "/c", "start", "", url],
        ["powershell.exe", "-NoProfile", "-Command", f"Start-Process '{url}'"],
    ]
    for command in commands:
        executable = shutil.which(command[0])
        if executable:
            subprocess.Popen(
                [executable, *command[1:]],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )
            return

    raise RuntimeError("No URL opener found on Windows.")


def windows_sendkeys_values(send_key: str) -> list[str]:
    normalized = send_key.strip().lower().replace(" ", "")
    if normalized == "auto":
        return ["^{ENTER}", "{ENTER}"]

    mappings = {
        "return": "{ENTER}",
        "enter": "{ENTER}",
        "ctrl+return": "^{ENTER}",
        "ctrl+enter": "^{ENTER}",
    }
    return [mappings.get(normalized, send_key)]


def powershell_string_array(values: list[str]) -> str:
    return "@(" + ", ".join("'" + value.replace("'", "''") + "'" for value in values) + ")"


def set_windows_clipboard_text(text: str) -> None:
    powershell = shutil.which("powershell.exe") or shutil.which("powershell")
    if not powershell:
        raise RuntimeError("powershell.exe was not found.")

    result = subprocess.run(
        [powershell, "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", "$input | Set-Clipboard"],
        input=text,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        detail = (result.stderr or result.stdout).strip()
        raise RuntimeError(detail or "Failed to write Teams message to the Windows clipboard.")


def paste_clipboard_into_teams_windows(timeout: float, send_keys: list[str] | None = None) -> bool:
    powershell = shutil.which("powershell.exe") or shutil.which("powershell")
    if not powershell:
        raise RuntimeError("powershell.exe was not found.")

    powershell_send_keys = powershell_string_array(send_keys or [])
    script = rf"""
Add-Type -AssemblyName UIAutomationClient
Add-Type -AssemblyName UIAutomationTypes
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing
Add-Type @"
using System;
using System.Runtime.InteropServices;
public class MouseClicker {{
    [DllImport("user32.dll")]
    public static extern bool SetCursorPos(int X, int Y);
    [DllImport("user32.dll")]
    public static extern void mouse_event(uint dwFlags, uint dx, uint dy, uint dwData, UIntPtr dwExtraInfo);
    public const uint MOUSEEVENTF_LEFTDOWN = 0x0002;
    public const uint MOUSEEVENTF_LEFTUP = 0x0004;
    public static void Click(int x, int y) {{
        SetCursorPos(x, y);
        mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, UIntPtr.Zero);
        mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, UIntPtr.Zero);
    }}
}}
"@

function Click-Center($element) {{
    try {{
        $rect = $element.Current.BoundingRectangle
        if ($rect.Width -gt 0 -and $rect.Height -gt 0) {{
            $x = [int]($rect.X + ($rect.Width / 2))
            $y = [int]($rect.Y + ($rect.Height / 2))
            [MouseClicker]::Click($x, $y)
            return $true
        }}
    }} catch {{}}
    return $false
}}

function Get-CandidateWindows($root) {{
    $windows = $root.FindAll([System.Windows.Automation.TreeScope]::Children, [System.Windows.Automation.Condition]::TrueCondition)
    $candidateWindows = @()
    foreach ($window in $windows) {{
        $name = $window.Current.Name
        if ($name -match 'Teams' -or $name -match 'Microsoft Edge' -or $name -match 'Google Chrome' -or $name -match 'Mozilla Firefox') {{
            $candidateWindows += $window
        }}
    }}
    return $candidateWindows
}}

$sendKeys = {powershell_send_keys}

function Try-FocusCompose($window) {{
    try {{
        $window.SetFocus()
    }} catch {{}}

    $windowRect = $window.Current.BoundingRectangle
    $conditions = @(
        [System.Windows.Automation.PropertyCondition]::new(
            [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
            [System.Windows.Automation.ControlType]::Edit
        ),
        [System.Windows.Automation.PropertyCondition]::new(
            [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
            [System.Windows.Automation.ControlType]::Document
        )
    )

    $composeCandidates = @()
    foreach ($condition in $conditions) {{
        $elements = $window.FindAll([System.Windows.Automation.TreeScope]::Descendants, $condition)
        foreach ($element in $elements) {{
            try {{
                if (-not $element.Current.IsEnabled -or $element.Current.IsOffscreen) {{
                    continue
                }}
                $rect = $element.Current.BoundingRectangle
                if ($rect.Width -le 40 -or $rect.Height -le 12) {{
                    continue
                }}
                $name = $element.Current.Name
                $lowerName = if ($name) {{ $name.ToLowerInvariant() }} else {{ "" }}
                $isBottomArea = $rect.Y -gt ($windowRect.Y + ($windowRect.Height * 0.55))
                $looksLikeCompose = (
                    $lowerName.Contains('message') -or
                    $lowerName.Contains('type') -or
                    $lowerName.Contains('compose') -or
                    $lowerName.Contains('chat') -or
                    $lowerName.Contains('메시지') -or
                    $lowerName.Contains('입력')
                )
                if ($isBottomArea -or $looksLikeCompose) {{
                    $composeCandidates += $element
                }}
            }} catch {{}}
        }}
    }}

    if ($composeCandidates.Count -gt 0) {{
        $target = $composeCandidates | Sort-Object {{ $_.Current.BoundingRectangle.Y }} | Select-Object -Last 1
        try {{
            $target.SetFocus()
        }} catch {{}}
        Click-Center $target | Out-Null
        Start-Sleep -Milliseconds 250
        [System.Windows.Forms.SendKeys]::SendWait('^v')
        Start-Sleep -Milliseconds 500
        foreach ($key in $sendKeys) {{
            [System.Windows.Forms.SendKeys]::SendWait($key)
            Start-Sleep -Milliseconds 700
        }}
        return $true
    }}

    try {{
        $x = [int]($windowRect.X + ($windowRect.Width * 0.50))
        $y = [int]($windowRect.Y + ($windowRect.Height - 95))
        [MouseClicker]::Click($x, $y)
        Start-Sleep -Milliseconds 250
        [System.Windows.Forms.SendKeys]::SendWait('^v')
        Start-Sleep -Milliseconds 500
        foreach ($key in $sendKeys) {{
            [System.Windows.Forms.SendKeys]::SendWait($key)
            Start-Sleep -Milliseconds 700
        }}
        return $true
    }} catch {{}}

    return $false
}}

$end = (Get-Date).AddSeconds({timeout})
$root = [System.Windows.Automation.AutomationElement]::RootElement

while ((Get-Date) -lt $end) {{
    foreach ($window in (Get-CandidateWindows $root)) {{
        if (Try-FocusCompose $window) {{
            exit 0
        }}
    }}
    Start-Sleep -Milliseconds 500
}}

exit 3
"""
    result = subprocess.run(
        [powershell, "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode == 0:
        return True
    if result.returncode == 3:
        return False

    detail = (result.stderr or result.stdout).strip()
    raise RuntimeError(detail or "Failed while trying to paste the Teams message.")


def click_teams_send_button_windows(timeout: float) -> bool:
    powershell = shutil.which("powershell.exe") or shutil.which("powershell")
    if not powershell:
        raise RuntimeError("powershell.exe was not found.")

    send_names = powershell_string_array(
        [
            "Send",
            "Send message",
            "Send now",
            "보내기",
            "메시지 보내기",
        ]
    )
    script = rf"""
Add-Type -AssemblyName UIAutomationClient
Add-Type -AssemblyName UIAutomationTypes
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing
Add-Type @"
using System;
using System.Runtime.InteropServices;
public class MouseClicker {{
    [DllImport("user32.dll")]
    public static extern bool SetCursorPos(int X, int Y);
    [DllImport("user32.dll")]
    public static extern void mouse_event(uint dwFlags, uint dx, uint dy, uint dwData, UIntPtr dwExtraInfo);
    public const uint MOUSEEVENTF_LEFTDOWN = 0x0002;
    public const uint MOUSEEVENTF_LEFTUP = 0x0004;
    public static void Click(int x, int y) {{
        SetCursorPos(x, y);
        mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, UIntPtr.Zero);
        mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, UIntPtr.Zero);
    }}
}}
"@

function Click-Element($element) {{
    try {{
        $element.SetFocus()
    }} catch {{}}

    try {{
        $pattern = $element.GetCurrentPattern([System.Windows.Automation.InvokePattern]::Pattern)
        $pattern.Invoke()
        return $true
    }} catch {{}}

    try {{
        $rect = $element.Current.BoundingRectangle
        if ($rect.Width -gt 0 -and $rect.Height -gt 0) {{
            $x = [int]($rect.X + ($rect.Width / 2))
            $y = [int]($rect.Y + ($rect.Height / 2))
            [MouseClicker]::Click($x, $y)
            return $true
        }}
    }} catch {{}}

    return $false
}}

function Get-CandidateWindows($root) {{
    $windows = $root.FindAll([System.Windows.Automation.TreeScope]::Children, [System.Windows.Automation.Condition]::TrueCondition)
    $candidateWindows = @()
    foreach ($window in $windows) {{
        $name = $window.Current.Name
        if ($name -match 'Teams' -or $name -match 'Microsoft Edge' -or $name -match 'Google Chrome' -or $name -match 'Mozilla Firefox') {{
            $candidateWindows += $window
        }}
    }}
    return $candidateWindows
}}

$sendNames = {send_names}
$end = (Get-Date).AddSeconds({timeout})
$root = [System.Windows.Automation.AutomationElement]::RootElement
$buttonCondition = [System.Windows.Automation.PropertyCondition]::new(
    [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
    [System.Windows.Automation.ControlType]::Button
)

while ((Get-Date) -lt $end) {{
    foreach ($window in (Get-CandidateWindows $root)) {{
        try {{
            $window.SetFocus()
        }} catch {{}}

        $windowRect = $window.Current.BoundingRectangle
        $buttons = $window.FindAll([System.Windows.Automation.TreeScope]::Descendants, $buttonCondition)
        $candidates = @()
        foreach ($button in $buttons) {{
            try {{
                if (-not $button.Current.IsEnabled -or $button.Current.IsOffscreen) {{
                    continue
                }}
                $rect = $button.Current.BoundingRectangle
                if ($rect.Width -le 0 -or $rect.Height -le 0) {{
                    continue
                }}
                if ($rect.Y -lt ($windowRect.Y + ($windowRect.Height * 0.55))) {{
                    continue
                }}
                $name = $button.Current.Name
                if ([string]::IsNullOrWhiteSpace($name)) {{
                    continue
                }}
                $lowerName = $name.ToLowerInvariant()
                if ($lowerName.Contains('file') -or $lowerName.Contains('attach') -or $lowerName.Contains('파일') -or $lowerName.Contains('첨부')) {{
                    continue
                }}
                foreach ($sendName in $sendNames) {{
                    $lowerSendName = $sendName.ToLowerInvariant()
                    if ($name -eq $sendName -or $lowerName -eq $lowerSendName -or $lowerName.Contains($lowerSendName)) {{
                        $candidates += $button
                        break
                    }}
                }}
            }} catch {{}}
        }}

        if ($candidates.Count -gt 0) {{
            $target = $candidates[$candidates.Count - 1]
            if (Click-Element $target) {{
                exit 0
            }}
        }}
    }}

    Start-Sleep -Milliseconds 500
}}

exit 3
"""
    result = subprocess.run(
        [powershell, "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        timeout=timeout + 15,
        check=False,
    )
    if result.returncode == 0:
        return True
    if result.returncode == 3:
        return False

    detail = (result.stderr or result.stdout).strip()
    raise RuntimeError(detail or "Failed while searching for the Teams Send button.")


def activate_teams_and_send_windows(send_key: str) -> None:
    powershell = shutil.which("powershell.exe") or shutil.which("powershell")
    if not powershell:
        raise RuntimeError("powershell.exe was not found.")

    keys = windows_sendkeys_values(send_key)
    powershell_keys = "@(" + ", ".join("'" + key.replace("'", "''") + "'" for key in keys) + ")"
    script = rf"""
$wshell = New-Object -ComObject WScript.Shell
$activated = $false
foreach ($title in @('Microsoft Teams', 'Teams', 'Chat', 'Google Chrome', 'Microsoft Edge', 'Mozilla Firefox')) {{
    if ($wshell.AppActivate($title)) {{
        $activated = $true
        break
    }}
}}
if (-not $activated) {{
    exit 2
}}
Start-Sleep -Milliseconds 500
Add-Type -AssemblyName System.Windows.Forms
foreach ($key in {powershell_keys}) {{
    [System.Windows.Forms.SendKeys]::SendWait($key)
    Start-Sleep -Milliseconds 700
}}
"""
    result = subprocess.run(
        [powershell, "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        detail = (result.stderr or result.stdout).strip()
        raise RuntimeError(detail or "Failed to activate Teams and send key on Windows.")


def send_teams_message_windows(send_method: str, send_key: str, teams_ready_timeout: float) -> None:
    if send_method in {"auto", "keys"}:
        activate_teams_and_send_windows(send_key)
        return

    if send_method == "button":
        clicked = click_teams_send_button_windows(teams_ready_timeout)
        if clicked:
            return
        raise RuntimeError("Could not find or click the Teams Send button.")


def prepare_teams_message_windows(
    *,
    message: str,
    input_method: str,
    paste_timeout: float,
    send_keys: list[str] | None = None,
) -> bool:
    if input_method == "deeplink":
        return False

    set_windows_clipboard_text(message)
    pasted = paste_clipboard_into_teams_windows(paste_timeout, send_keys=send_keys)
    if not pasted and input_method == "clipboard":
        raise RuntimeError("Could not paste the Teams message into the compose box.")
    return pasted


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crawl ICS password and send it to your Teams self-chat on Windows."
    )
    parser.add_argument(
        "--sso-id",
        default=os.environ.get("ICS_SSO_ID"),
        help="SSO ID. Defaults to ICS_SSO_ID. Used for keyring lookup when ICS_SSO_PW is unset.",
    )
    parser.add_argument(
        "--keyring-service",
        default=os.environ.get("ICS_KEYRING_SERVICE", DEFAULT_KEYRING_SERVICE),
        help="keyring service name for SSO password lookup. Defaults to ICS_SSO.",
    )
    parser.add_argument(
        "--teams-user",
        default=os.environ.get("TEAMS_SELF_UPN"),
        help="Teams login email/UPN. Defaults to TEAMS_SELF_UPN.",
    )
    parser.add_argument(
        "--tenant-id",
        default=os.environ.get("TEAMS_TENANT_ID"),
        help="Optional Microsoft Entra tenant ID. Defaults to TEAMS_TENANT_ID.",
    )
    parser.add_argument("--config-excel", default=DEFAULT_CONFIG_EXCEL)
    parser.add_argument(
        "--browser",
        choices=["auto", "edge", "chrome"],
        default=os.environ.get("ICS_BROWSER", "edge"),
        help="Browser for ICS crawling. Defaults to edge.",
    )
    parser.add_argument("--ics-headless", action="store_true", help="Run the ICS browser headless.")
    parser.add_argument(
        "--keep-ics-browser-open",
        action="store_true",
        help="Leave the ICS Selenium browser open after crawling.",
    )
    parser.add_argument("--cookie-timeout", type=float, default=DEFAULT_COOKIE_TIMEOUT)
    parser.add_argument(
        "--login-wait",
        type=float,
        default=DEFAULT_LOGIN_WAIT,
        help="Seconds to wait after clicking the portal login button before entering ICS SSO.",
    )
    parser.add_argument(
        "--ics-retries",
        type=int,
        default=DEFAULT_ICS_RETRIES,
        help="Retry count for the ICS SSO URL and icsPwd cookie lookup.",
    )
    parser.add_argument(
        "--ics-retry-delay",
        type=float,
        default=DEFAULT_ICS_RETRY_DELAY,
        help="Seconds to wait between ICS SSO retries.",
    )
    parser.add_argument(
        "--debug-dir",
        help="Optional directory for current URL/title and screenshots on ICS SSO failures.",
    )
    parser.add_argument(
        "--message-prefix",
        default=os.environ.get("TEAMS_MESSAGE_PREFIX", ""),
        help="Optional prefix before the crawled password.",
    )
    parser.add_argument(
        "--no-open",
        action="store_true",
        help="Print the Teams deep link instead of opening it.",
    )
    parser.add_argument(
        "--no-send",
        action="store_true",
        help="Open Teams and pre-fill the message, but do not press Send.",
    )
    parser.add_argument("--send-delay", type=float, default=DEFAULT_SEND_DELAY)
    parser.add_argument(
        "--teams-input-method",
        choices=["clipboard", "deeplink"],
        default=os.environ.get("TEAMS_INPUT_METHOD", DEFAULT_TEAMS_INPUT_METHOD),
        help="How to put text into Teams. clipboard opens chat, pastes text, then sends.",
    )
    parser.add_argument(
        "--send-method",
        choices=["auto", "button", "keys"],
        default=os.environ.get("TEAMS_SEND_METHOD", DEFAULT_SEND_METHOD),
        help="Windows send strategy. keys is recommended after clipboard paste.",
    )
    parser.add_argument(
        "--send-key",
        default=os.environ.get("TEAMS_SEND_KEY", DEFAULT_SEND_KEY),
        help="Send key. Defaults to auto, which tries Return and then ctrl+Return.",
    )
    parser.add_argument(
        "--teams-ready-timeout",
        type=float,
        default=DEFAULT_TEAMS_READY_TIMEOUT,
        help="Seconds to wait for Teams compose/send controls before falling back.",
    )
    parser.add_argument("--max-url-chars", type=int, default=DEFAULT_MAX_URL_CHARS)
    parser.add_argument(
        "--print-password",
        action="store_true",
        help="Also print the crawled password to stdout. Disabled by default.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not args.teams_user:
        print(
            "Missing Teams user. Pass --teams-user your.name@company.com "
            "or set TEAMS_SELF_UPN.",
            file=sys.stderr,
        )
        return 2

    browser = choose_browser(args.browser)
    sso_id, sso_pw = read_sso_credentials(
        config_excel=args.config_excel,
        sso_id=args.sso_id,
        keyring_service=args.keyring_service,
    )
    ics_password = crawl_ics_password(
        sso_id=sso_id,
        sso_pw=sso_pw,
        browser=browser,
        headless=args.ics_headless,
        keep_browser_open=args.keep_ics_browser_open,
        cookie_timeout=args.cookie_timeout,
        login_wait=args.login_wait,
        ics_retries=args.ics_retries,
        ics_retry_delay=args.ics_retry_delay,
        debug_dir=args.debug_dir,
    )

    if args.print_password:
        print(ics_password)
    else:
        print(f"ICS password crawled. length={len(ics_password)}")

    if not ics_password.strip():
        raise RuntimeError("ICS password was empty; refusing to open or send an empty Teams message.")

    message = f"{args.message_prefix}{ics_password}"
    url = build_teams_chat_deep_link(
        user=args.teams_user,
        message=message if args.teams_input_method == "deeplink" else None,
        tenant_id=args.tenant_id,
    )

    if len(url) > args.max_url_chars:
        print(
            f"Teams deep link is too long: {len(url)} chars. "
            "Use a shorter message prefix or increase --max-url-chars.",
            file=sys.stderr,
        )
        return 3

    if args.no_open:
        print(url)
        return 0

    open_url(url)
    time.sleep(args.send_delay)

    if args.no_send:
        prepare_teams_message_windows(
            message=message,
            input_method=args.teams_input_method,
            paste_timeout=args.teams_ready_timeout,
        )
        print("Opened Teams chat and prepared the message; --no-send left it unsent.")
        return 0

    if args.teams_input_method == "clipboard" and args.send_method in {"auto", "keys"}:
        prepare_teams_message_windows(
            message=message,
            input_method=args.teams_input_method,
            paste_timeout=args.teams_ready_timeout,
            send_keys=windows_sendkeys_values(args.send_key),
        )
        print("Pasted ICS password and sent Teams key sequence.")
        return 0

    prepare_teams_message_windows(
        message=message,
        input_method=args.teams_input_method,
        paste_timeout=args.teams_ready_timeout,
    )
    send_teams_message_windows(args.send_method, args.send_key, args.teams_ready_timeout)
    print("Sent Teams send action after preparing the message.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
